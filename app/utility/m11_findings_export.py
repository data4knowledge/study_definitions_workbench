"""Server-side formatters for M11 validation findings.

Each ``to_*`` helper takes a list of finding dicts (the shape emitted
by ``M11ValidationResults.to_dict()``) and returns a bytes / string
payload ready to stream as an HTTP response. Used by the
``/validate/m11-docx/download/{csv,json,md,xlsx}`` routes — one
helper per format keeps route code trivial and puts format logic
in one place.

Why server-side? Earlier iterations generated CSV / JSON / Markdown
in the browser via JS + Blob. Zero-JS became a project direction
(see docs/lessons_learned.md), so the generators moved here.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date


_FIELDS = (
    "rule_id",
    "severity",
    "status",
    "element_name",
    "section_title",
    "message",
    "expected",
    "actual",
)

_MD_HEADERS = (
    "Rule",
    "Severity",
    "Status",
    "Element",
    "Section",
    "Message",
    "Expected",
    "Actual",
)


def default_filename(source_filename: str, extension: str) -> str:
    """Build a deterministic filename of the form
    ``{source-basename}-m11-findings-{YYYY-MM-DD}.{extension}``.

    Callers can pass it through a query-string override; this is the
    fallback when none is provided.
    """
    base = (source_filename or "protocol").rsplit(".", 1)[0]
    return f"{base}-m11-findings-{date.today().isoformat()}.{extension}"


def sanitise_filename(name: str, fallback: str) -> str:
    """Restrict to a safe subset so the name is inert in a
    Content-Disposition header. Returns the fallback on empty input."""
    cleaned = "".join(c for c in (name or "") if c.isalnum() or c in "-_. ")
    return cleaned or fallback


def to_csv(findings: list[dict]) -> bytes:
    """Serialise findings as UTF-8 CSV with a header row. RFC-4180
    escaping applies (DictWriter handles commas / quotes / newlines).
    """
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(_FIELDS), extrasaction="ignore")
    writer.writeheader()
    for finding in findings or []:
        row = {field: _stringify(finding.get(field)) for field in _FIELDS}
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def to_json(findings: list[dict]) -> bytes:
    """Pretty-printed JSON. Preserves the field order the server
    serialises so diffing two reports is straightforward."""
    return json.dumps(findings or [], indent=2).encode("utf-8")


def to_markdown(findings: list[dict], source_filename: str) -> bytes:
    """A readable Markdown table with a heading and metadata. Useful for
    pasting into PRs, tickets, emails, or chat."""
    lines: list[str] = []
    lines.append("# M11 Validation Findings")
    lines.append("")
    lines.append(f"**Source:** {source_filename or '(unknown)'}")
    lines.append(f"**Generated:** {date.today().isoformat()}")
    lines.append(f"**Findings:** {len(findings or [])}")
    lines.append("")
    if not findings:
        lines.append("_No findings._")
        return ("\n".join(lines) + "\n").encode("utf-8")
    lines.append("| " + " | ".join(_MD_HEADERS) + " |")
    lines.append("| " + " | ".join("---" for _ in _MD_HEADERS) + " |")
    for finding in findings:
        cells = [_md_cell(finding.get(field)) for field in _FIELDS]
        lines.append("| " + " | ".join(cells) + " |")
    return ("\n".join(lines) + "\n").encode("utf-8")


def to_xlsx(findings: list[dict], source_filename: str) -> io.BytesIO:
    """Render an openpyxl Workbook with two sheets: the findings
    table on the first sheet, and source-filename / generation-date
    metadata on the second. Returns an ``io.BytesIO`` ready to stream.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "M11 Findings"

    headers = [
        ("Rule", 10),
        ("Severity", 10),
        ("Status", 10),
        ("Element", 28),
        ("Section", 18),
        ("Message", 60),
        ("Expected", 30),
        ("Actual", 30),
    ]
    for col_idx, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, finding in enumerate(findings or [], start=2):
        for col_idx, field in enumerate(_FIELDS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_stringify(finding.get(field)))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    meta = wb.create_sheet("Source")
    meta["A1"] = "Source file"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = source_filename or ""
    meta["A3"] = "Generated"
    meta["A3"].font = Font(bold=True)
    meta["A4"] = date.today().isoformat()
    meta.column_dimensions["A"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- internals -----------------------------------------------------


def _stringify(value) -> str:
    """Coerce None / non-str values to a plain string so formatters
    don't have to branch."""
    if value is None:
        return ""
    return str(value)


def _md_cell(value) -> str:
    """Escape pipes and newlines for a markdown table cell."""
    return _stringify(value).replace("|", "\\|").replace("\n", " ")
