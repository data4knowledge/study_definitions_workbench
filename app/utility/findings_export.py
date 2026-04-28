"""Server-side formatters for validation findings (engine-agnostic).

Each ``to_*`` helper takes a list of finding dicts and returns a bytes /
string payload ready to stream as an HTTP response. Used by the shared
``/validate/download/{csv,json,md,xlsx}`` routes — one helper per
format keeps route code trivial and puts format logic in one place.

Findings here are produced by any of SDW's validation flows:
M11 docx (``M11Validator``), USDM v4 rules (``USDM4.validate``), and
USDM v4 CORE (``USDM4.validate_core``). The formatters don't care
which — each call site hands over a uniform row shape plus two
presentation hints (``kind`` and ``title``).

Row shape
---------

Authoritative row shape is the one produced by
:mod:`app.utility.finding_projections`:

    rule_id    — stable rule identifier
    severity   — "error" | "warning" | "info"
    section    — where in the document the finding applies
    element    — the specific element or attribute
    message    — human-readable finding text
    rule_text  — optional rule description (may be empty)
    path       — JSON/DOCX path for drill-down (may be empty)

Why server-side? Earlier iterations generated CSV / JSON / Markdown
in the browser via JS + Blob. Zero-JS became a project direction
(see docs/lessons_learned.md), so the generators moved here.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date


# The canonical field set — mirrors the column set rendered in the
# shared ``validate/partials/results.html`` table so the download
# carries every field the user sees on screen.  The display fuses
# ``section`` and ``element`` into a single two-line "Element" column
# for layout reasons; the download keeps them as separate columns
# because spreadsheet analysis benefits from granular columns for
# sort / filter / pivot.  Order tracks the display: location first
# (Section, Element), then the categorical (Severity), the rule
# identity (Rule + Description), the actual finding text (Message),
# and finally the drill-down JSON path.
_FIELDS = (
    "section",
    "element",
    "severity",
    "rule_id",
    "rule_text",
    "message",
    "path",
)

# Header label per field, used by the markdown / xlsx formatters and
# the CSV writer.  Single-source-of-truth so adding or renaming a
# column is a one-line change.
_HEADER_FOR = {
    "section": "Section",
    "element": "Element",
    "severity": "Severity",
    "rule_id": "Rule",
    "rule_text": "Description",
    "message": "Message",
    "path": "Path",
}

# XLSX column width per field.  Picked to match the relative
# information density of each column — Section / Element are tag-like,
# Severity is a short word, Rule is a short id, Description and
# Message are prose paragraphs (Message tends to be longest), Path is
# a long-string JSONPath / DOCX path.
_WIDTH_FOR = {
    "section": 24,
    "element": 28,
    "severity": 10,
    "rule_id": 14,
    "rule_text": 50,
    "message": 70,
    "path": 50,
}


def _active_fields(findings: list[dict]) -> tuple[str, ...]:
    """Return ``_FIELDS`` minus columns that carry no information for
    this particular set of findings.  Today only ``path`` is
    data-driven: M11 findings don't surface drill-down paths, so the
    column would otherwise be a wasted stripe in every download.
    Mirrors the column-visibility rule on the display side
    (``validate/partials/results.html``) so what the user sees on
    screen and what they get in the download stay in lock-step.

    Empty findings list defaults to suppressing ``path`` — there's
    nothing to suggest paths would be useful, and the M11 case is
    the most common reason to download empty results.
    """
    if findings and any(
        (f.get("path") if isinstance(f, dict) else "") for f in findings
    ):
        return _FIELDS
    return tuple(f for f in _FIELDS if f != "path")


def default_filename(
    source_filename: str,
    extension: str,
    kind: str = "validation-findings",
) -> str:
    """Build a deterministic filename of the form
    ``{source-basename}-{kind}-{YYYY-MM-DD}.{extension}``.

    ``kind`` is the validator tag (e.g. ``"m11-findings"``,
    ``"usdm-cdisc-findings"``, ``"usdm-d4k-findings"``). Callers can
    pass any slug-safe string; the default is intentionally generic so
    call sites that don't care still produce a reasonable name.
    """
    base = (source_filename or "protocol").rsplit(".", 1)[0]
    return f"{base}-{kind}-{date.today().isoformat()}.{extension}"


def sanitise_filename(name: str, fallback: str) -> str:
    """Restrict to a safe subset so the name is inert in a
    Content-Disposition header. Returns the fallback on empty input."""
    cleaned = "".join(c for c in (name or "") if c.isalnum() or c in "-_. ")
    return cleaned or fallback


def to_csv(findings: list[dict]) -> bytes:
    """Serialise findings as UTF-8 CSV with a header row. RFC-4180
    escaping applies (DictWriter handles commas / quotes / newlines).
    """
    fields = _active_fields(findings)
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=list(fields), extrasaction="ignore")
    writer.writeheader()
    for finding in findings or []:
        row = _row_view(finding)
        writer.writerow({field: _stringify(row.get(field)) for field in fields})
    return buffer.getvalue().encode("utf-8")


def to_json(findings: list[dict]) -> bytes:
    """Pretty-printed JSON. Normalised to the canonical row shape so
    downstream consumers (diffing, archiving) see a stable key set
    regardless of which engine produced the findings."""
    fields = _active_fields(findings)
    normalised = [
        {field: _stringify(_row_view(finding).get(field)) for field in fields}
        for finding in findings or []
    ]
    return json.dumps(normalised, indent=2).encode("utf-8")


def to_markdown(
    findings: list[dict],
    source_filename: str,
    title: str = "Validation Findings",
) -> bytes:
    """A readable Markdown table with a heading and metadata. Useful for
    pasting into PRs, tickets, emails, or chat.

    ``title`` varies by validator (e.g. ``"M11 Validation Findings"``,
    ``"USDM v4 d4k Findings"``, ``"USDM v4 CDISC Findings"``).
    """
    lines: list[str] = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append(f"**Source:** {source_filename or '(unknown)'}")
    lines.append(f"**Generated:** {date.today().isoformat()}")
    lines.append(f"**Findings:** {len(findings or [])}")
    lines.append("")
    if not findings:
        lines.append("_No findings._")
        return ("\n".join(lines) + "\n").encode("utf-8")
    fields = _active_fields(findings)
    headers = [_HEADER_FOR[f] for f in fields]
    lines.append("| " + " | ".join(headers) + " |")
    lines.append("| " + " | ".join("---" for _ in headers) + " |")
    for finding in findings:
        row = _row_view(finding)
        cells = [_md_cell(row.get(field)) for field in fields]
        lines.append("| " + " | ".join(cells) + " |")
    return ("\n".join(lines) + "\n").encode("utf-8")


def to_xlsx(
    findings: list[dict],
    source_filename: str,
    sheet_title: str = "Findings",
) -> io.BytesIO:
    """Render an openpyxl Workbook with two sheets: the findings
    table on the first sheet, and source-filename / generation-date
    metadata on the second. Returns an ``io.BytesIO`` ready to stream.

    ``sheet_title`` names the findings worksheet — e.g. ``"M11 Findings"``,
    ``"USDM d4k Findings"``, ``"USDM CDISC Findings"``. Excel caps
    worksheet titles at 31 characters; callers should stay within that.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel hard limit.

    # Active column set is data-driven (see ``_active_fields``) so the
    # workbook tracks the display: Path is dropped when no finding
    # carries a non-empty path.  Header label and width come from the
    # ``_HEADER_FOR`` / ``_WIDTH_FOR`` maps so the three formatters
    # share one source of truth.
    fields = _active_fields(findings)
    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_HEADER_FOR[field])
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = _WIDTH_FOR[field]

    for row_idx, finding in enumerate(findings or [], start=2):
        row = _row_view(finding)
        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(
                row=row_idx, column=col_idx, value=_stringify(row.get(field))
            )
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    meta = wb.create_sheet("Source")
    meta["A1"] = "Source file"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = source_filename or ""
    meta["A3"] = "Generated"
    meta["A3"].font = Font(bold=True)
    meta["A4"] = date.today().isoformat()
    meta.column_dimensions["A"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# --- internals -----------------------------------------------------


def _row_view(finding: dict) -> dict[str, str]:
    """Normalise one finding dict to the canonical row shape produced
    by :mod:`app.utility.finding_projections`. Non-dict inputs collapse
    to all-empty values rather than raising, so a malformed row from a
    runaway validator does not kill the download."""
    if not isinstance(finding, dict):
        return {field: "" for field in _FIELDS}
    return {
        "section": finding.get("section") or "",
        "element": finding.get("element") or "",
        "severity": finding.get("severity") or "",
        "rule_id": finding.get("rule_id") or "",
        "rule_text": finding.get("rule_text") or "",
        "message": finding.get("message") or "",
        "path": finding.get("path") or "",
    }


def _stringify(value) -> str:
    """Coerce None / non-str values to a plain string so formatters
    don't have to branch."""
    if value is None:
        return ""
    return str(value)


def _md_cell(value) -> str:
    """Escape pipes and newlines for a markdown table cell."""
    return _stringify(value).replace("|", "\\|").replace("\n", " ")
