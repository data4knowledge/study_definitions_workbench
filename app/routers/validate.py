import io
from datetime import date
from fastapi import APIRouter, Depends, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.database import get_db
from app.dependencies.dependency import protect_endpoint
from app.dependencies.utility import user_details
from app.dependencies.templates import templates
from app.configuration.configuration import application_configuration
from app.model.file_handling.local_files import LocalFiles
from app.model.file_handling.data_files import DataFiles
from usdm_info import __model_version__ as usdm_version
from app.database.user import User

from app.imports.form_handler import FormHandler
from app.utility.m11_annotate import annotate as m11_annotate
from usdm4 import USDM4, RulesValidationResults
from usdm3 import USDM3
from usdm4_protocol.m11 import USDM4M11

router = APIRouter(
    prefix="/validate", tags=["validate"], dependencies=[Depends(protect_endpoint)]
)


@router.get("/usdm3", dependencies=[Depends(protect_endpoint)])
def validate_usdm3(request: Request, session: Session = Depends(get_db)):
    return _validate_setup(
        request,
        session,
        "json",
        False,
        "/validate/usdm3",
        "validate/partials/validate_json.html",
        {"version": "3.0.0"},
    )


@router.get("/usdm", dependencies=[Depends(protect_endpoint)])
def validate_usdm(request: Request, session: Session = Depends(get_db)):
    return _validate_setup(
        request,
        session,
        "json",
        False,
        "/validate/usdm",
        "validate/partials/validate_json.html",
        {"version": usdm_version},
    )


@router.post("/usdm3", dependencies=[Depends(protect_endpoint)])
async def validate_usdm3_process(
    request: Request, source: str = "browser", session: Session = Depends(get_db)
):
    user, present_in_db = user_details(request, session)
    return await _process(request, user, USDM3(), source)


@router.post("/usdm", dependencies=[Depends(protect_endpoint)])
async def validate_usdm_process(
    request: Request, source: str = "browser", session: Session = Depends(get_db)
):
    user, present_in_db = user_details(request, session)
    return await _process(request, user, USDM4(), source)


@router.get("/m11-docx", dependencies=[Depends(protect_endpoint)])
def validate_m11_docx(request: Request, session: Session = Depends(get_db)):
    """File picker for M11 ``.docx`` validation. Extracts the protocol via
    USDM4M11 then runs the M11 specification checks."""
    return _validate_setup(
        request,
        session,
        "docx",
        False,
        "/validate/m11-docx",
        "validate/partials/validate_m11_docx.html",
        {"version": "ICH M11 (2025-11-16)"},
    )


@router.post("/m11-docx", dependencies=[Depends(protect_endpoint)])
async def validate_m11_docx_process(
    request: Request, source: str = "browser", session: Session = Depends(get_db)
):
    user, present_in_db = user_details(request, session)
    return await _process_m11_docx(request, user, source)


@staticmethod
def _validate_setup(
    request: Request,
    session: Session,
    ext: str,
    other_file: bool,
    data_url: str,
    page_html: str,
    extra_data: dict = {},
):
    user, present_in_db = user_details(request, session)
    data = application_configuration.file_picker
    data.update(extra_data)
    data["dir"] = LocalFiles().root if data["os"] else ""
    data["required_ext"] = ext
    data["other_files"] = other_file
    data["url"] = data_url
    return templates.TemplateResponse(request, page_html, {"user": user, "data": data})


@staticmethod
async def _process(request: Request, user: User, usdm: USDM3 | USDM4, source: str):
    form_handler = FormHandler(
        request,
        False,
        ".json",
        source,
    )
    main_file, image_files, messages = await form_handler.get_files()
    # print(f"MAIN FILE: {main_file['filename']}")
    if main_file:
        files = DataFiles()
        _ = files.new()
        files.save("usdm", main_file["contents"], main_file["filename"])
        full_path, filename, exists = files.path("usdm")
        results: RulesValidationResults = usdm.validate(full_path)
        errors = results.to_dict()
        files.delete()
    else:
        print(f"Messages: {messages}")
        errors = []
        messages.append("Failed to process the validation file")
    return templates.TemplateResponse(
        request,
        "validate/partials/results.html",
        {
            "user": user,
            "data": {
                "filename": main_file,
                "messages": messages,
                "errors": errors,
            },
        },
    )


@staticmethod
async def _process_m11_docx(request: Request, user: User, source: str):
    """Process an uploaded M11 ``.docx`` through USDM4M11 validation and
    render the findings template.

    Mirrors ``_process`` for USDM JSON. Differences: ``.docx`` extension,
    USDM4M11 entry point, and a different results partial because M11
    findings carry element + section location rather than klass/attribute.
    """
    form_handler = FormHandler(
        request,
        False,
        ".docx",
        source,
    )
    main_file, image_files, messages = await form_handler.get_files()
    findings: list[dict] = []
    annotated_html: str = ""
    unplaced_findings: list[dict] = []
    if main_file:
        files = DataFiles()
        _ = files.new()
        # Reuse the existing "docx" media type; "m11" isn't registered in
        # DataFiles.media_type and shouldn't be added just for validation.
        files.save("docx", main_file["contents"], main_file["filename"])
        full_path, filename, exists = files.path("docx")
        m11 = USDM4M11()
        results = m11.validate_docx(full_path)
        if results is None:
            messages.append("Validation could not be completed (extraction failed).")
        else:
            findings = results.to_dict()
            # Render the extracted wrapper as M11 HTML and overlay the
            # findings so the "Annotated document" tab on the results
            # page shows each finding at the element it applies to.
            # render_current reuses the wrapper cached during
            # validate_docx — no second extraction.
            rendered = m11.render_current() or ""
            annotated = m11_annotate(rendered, findings)
            annotated_html = annotated.html
            unplaced_findings = annotated.unplaced
        files.delete()
    else:
        messages.append("Failed to process the validation file")
    return templates.TemplateResponse(
        request,
        "validate/partials/m11_docx_results.html",
        {
            "user": user,
            "data": {
                "filename": main_file,
                "messages": messages,
                "findings": findings,
                "annotated_html": annotated_html,
                "unplaced_findings": unplaced_findings,
            },
        },
    )


@router.post("/m11-docx/download/xlsx", dependencies=[Depends(protect_endpoint)])
async def validate_m11_docx_download_xlsx(request: Request):
    """Stream findings as an .xlsx. Client POSTs the findings JSON it
    already has (embedded in the results page) plus the source filename;
    the server round-trip exists only because we don't ship a
    client-side XLSX generator. CSV / JSON / Markdown are produced
    entirely in the browser.

    The ``filename`` query-string lets the caller choose the download
    filename; we default to something sensible if it's absent or
    malformed. The file is generated via openpyxl (already a transitive
    dependency through usdm4_excel).
    """
    payload = await request.json()
    findings = payload.get("findings") or []
    source_filename = payload.get("source_filename") or "protocol.docx"
    default_name = (
        f"{source_filename.rsplit('.', 1)[0]}"
        f"-m11-findings-{date.today().isoformat()}.xlsx"
    )
    filename = request.query_params.get("filename") or default_name
    # Only allow a narrow filename shape so the Content-Disposition
    # header stays safe from injection.
    filename = "".join(
        c for c in filename if c.isalnum() or c in "-_. "
    ) or default_name

    buffer = _findings_to_xlsx(findings, source_filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


def _findings_to_xlsx(findings: list[dict], source_filename: str) -> io.BytesIO:
    """Render a list of finding dicts to an in-memory .xlsx. Uses
    openpyxl (installed via usdm4_excel). Bold header row, column
    widths sized to typical finding content so the sheet opens
    readable in Excel / Numbers / Sheets without manual resizing.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "M11 Findings"

    headers = [
        ("Rule", 10),
        ("Severity", 10),
        ("Status", 10),
        ("Element", 28),
        ("Section", 18),
        ("Message", 60),
        ("Expected", 30),
        ("Actual", 30),
    ]
    for col_idx, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, f in enumerate(findings, start=2):
        values = [
            f.get("rule_id", ""),
            f.get("severity", ""),
            f.get("status", ""),
            f.get("element_name", ""),
            f.get("section_title", ""),
            f.get("message", ""),
            f.get("expected", ""),
            f.get("actual", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Source-filename note on a second sheet so the primary sheet
    # stays dense. Useful when someone sends the xlsx around and loses
    # the original filename context.
    meta = wb.create_sheet("Source")
    meta["A1"] = "Source file"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = source_filename
    meta["A3"] = "Generated"
    meta["A3"].font = Font(bold=True)
    meta["A4"] = date.today().isoformat()
    meta.column_dimensions["A"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
